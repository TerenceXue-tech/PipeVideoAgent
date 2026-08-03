#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

VIDEO_EXTENSIONS = {
    ".mp4",
    ".avi",
    ".mov",
    ".mkv",
    ".flv",
    ".wmv",
    ".m4v",
    ".ts",
    ".webm",
    ".mpeg",
    ".mpg",
}
TIMESTAMP_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}")


def collect_videos(root_dir: Path) -> list[Path]:
    videos = [
        p for p in root_dir.rglob("*") if p.is_file() and p.suffix.lower() in VIDEO_EXTENSIONS
    ]
    return sorted(videos)


def extract_timestamp(text: str) -> str | None:
    match = TIMESTAMP_PATTERN.search(text)
    return match.group(0) if match else None


def load_timestamps_from_excel(excel_path: Path) -> list[str]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active
    timestamps: list[str] = []
    invalid_rows: list[tuple[int, str]] = []

    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True), start=1
    ):
        cell_value = row[0] if row else None
        if cell_value is None:
            continue

        raw_text = str(cell_value).strip()
        if not raw_text:
            continue

        # Skip common header texts in the first row.
        if row_idx == 1 and ("时间戳" in raw_text.lower() or "timestamp" in raw_text.lower()):
            continue

        timestamp = extract_timestamp(raw_text)
        if timestamp is None:
            invalid_rows.append((row_idx, raw_text))
            continue
        timestamps.append(timestamp)

    if invalid_rows:
        print(f"[WARN] Ignored {len(invalid_rows)} rows with invalid timestamp format.")
        for row_idx, raw_text in invalid_rows[:10]:
            print(f"  - row {row_idx}: {raw_text}")
        if len(invalid_rows) > 10:
            print(f"  ... and {len(invalid_rows) - 10} more rows.")

    return timestamps


def build_video_index(videos: list[Path]) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    for video_path in videos:
        timestamp = extract_timestamp(video_path.stem)
        if timestamp is None:
            continue
        index.setdefault(timestamp, []).append(video_path)

    for timestamp in index:
        index[timestamp].sort()
    return index


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Read timestamps from the first column of an Excel file and create "
            "video_list.json by matching files under video_data."
        )
    )
    parser.add_argument(
        "--excel-path",
        type=Path,
        default=Path("/home/xtc/PipeVideo/labeled_data.xlsx"),
        help="Excel file path. Timestamps are read from the first column.",
    )
    parser.add_argument(
        "--video-data-dir",
        type=Path,
        default=Path("/home/xtc/PipeVideo/video_data"),
        help="Root directory that contains videos.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("/home/xtc/PipeVideo/video_list.json"),
        help="Output JSON file path.",
    )
    args = parser.parse_args()

    if not args.excel_path.exists():
        raise FileNotFoundError(f"Excel file does not exist: {args.excel_path}")
    if not args.video_data_dir.exists():
        raise FileNotFoundError(f"Video data directory does not exist: {args.video_data_dir}")

    timestamps = load_timestamps_from_excel(args.excel_path)
    if not timestamps:
        raise RuntimeError(f"No valid timestamps found in the first column: {args.excel_path}")

    all_videos = collect_videos(args.video_data_dir)
    if not all_videos:
        raise RuntimeError(f"No video files found under: {args.video_data_dir}")

    video_index = build_video_index(all_videos)
    payload: list[dict[str, str]] = []
    missing_timestamps: list[str] = []
    duplicate_matches: dict[str, list[Path]] = {}

    for timestamp in timestamps:
        matches = video_index.get(timestamp, [])
        if not matches:
            missing_timestamps.append(timestamp)
            continue
        if len(matches) > 1:
            duplicate_matches[timestamp] = matches
        payload.append({"video_path": str(matches[0])})

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Read {len(timestamps)} timestamps from: {args.excel_path}")
    print(f"Indexed {len(all_videos)} videos under: {args.video_data_dir}")
    print(f"Matched {len(payload)} videos and wrote: {args.output}")

    if missing_timestamps:
        print(f"[WARN] {len(missing_timestamps)} timestamps were not found in video_data.")
        for timestamp in missing_timestamps[:10]:
            print(f"  - missing: {timestamp}")
        if len(missing_timestamps) > 10:
            print(f"  ... and {len(missing_timestamps) - 10} more missing timestamps.")

    if duplicate_matches:
        print(f"[WARN] {len(duplicate_matches)} timestamps matched multiple files. Kept first sorted.")
        for timestamp, paths in list(duplicate_matches.items())[:10]:
            print(f"  - {timestamp}: {[str(path) for path in paths]}")
        if len(duplicate_matches) > 10:
            print(f"  ... and {len(duplicate_matches) - 10} more duplicate timestamps.")


if __name__ == "__main__":
    main()
